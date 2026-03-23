#!/usr/bin/env python3
"""
map-tasks-to-issues - Skill para mapear tasks de requirements con issues de GitHub

Uso: python3 map_tasks_to_issues.py <repository_name>
Ejemplo: python3 map_tasks_to_issues.py goetz-kundenportal-phoenix
"""

import json
import sys
import os
from datetime import datetime, timezone
from pathlib import Path
from difflib import SequenceMatcher

# Intentar importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  Advertencia: openpyxl no está instalado.")
    print("   Intenta: pip3 install openpyxl")
    print("   Continuando con fallback a CSV...")
    import csv

class TaskIssueMapper:
    def __init__(self, repository_name):
        self.repo_name = repository_name
        self.base_path = Path(f"/Users/admin/dev/Reports/{repository_name}")
        self.requirements_file = self.base_path / "requirements.json"
        self.board_issues_file = self.base_path / "board-issues.json"
        
    def validate_files(self):
        """Validar que los archivos requeridos existan"""
        if not self.base_path.exists():
            print(f"❌ Error: Directorio no encontrado: {self.base_path}")
            sys.exit(1)
        
        if not self.requirements_file.exists():
            print(f"❌ Error: Archivo no encontrado: {self.requirements_file}")
            sys.exit(1)
            
        if not self.board_issues_file.exists():
            print(f"❌ Error: Archivo no encontrado: {self.board_issues_file}")
            sys.exit(1)
        
        return True
    
    def load_json(self, file_path):
        """Cargar archivo JSON"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"❌ Error al cargar {file_path}: {str(e)}")
            sys.exit(1)
    
    def extract_tasks(self, requirements):
        """Extraer tasks de los requirements"""
        tasks = []
        for req in requirements.get('main_requirements', []):
            req_id = req['id']
            req_title = req['title']
            
            for task in req.get('sub_tasks', []):
                tasks.append({
                    'id': task['id'],
                    'title': task['title'],
                    'description': task.get('description', ''),
                    'requirement_id': req_id,
                    'requirement_title': req_title
                })
        
        return tasks
    
    def extract_issues(self, board_issues):
        """Extraer issues del board"""
        issues = []
        for issue in board_issues.get('issues', []):
            issues.append({
                'number': issue['number'],
                'title': issue['title'],
                'body': issue.get('body', ''),
                'state': issue['state'],
                'created_at': issue.get('created_at', ''),
                'updated_at': issue.get('updated_at', '')
            })
        
        return issues
    
    def normalize_text(self, text):
        """Normalizar texto para comparación"""
        if not text:
            return ""
        
        # Convertir a minúsculas
        text = text.lower()
        
        # Mapeo de palabras clave alemanas a conceptos
        # Para mejor similitud con textos en inglés
        replacements = {
            'formulare': 'form',
            'formular': 'form',
            'tabelle': 'table',
            'datenstruktur': 'datastructure',
            'struktur': 'structure',
            'bearbeiter': 'editor',
            'editor': 'editor',
            'version': 'version',
            'draft': 'draft',
            'entwurf': 'draft',
            'veroffentlichen': 'publish',
            'erstellen': 'create',
            'löschen': 'delete',
            'bearbeitbar': 'editable',
            'verwalten': 'manage',
            'verwaltungs': 'management',
            'oberfläche': 'interface',
            'seeding': 'seeding',
            'grpc': 'grpc',
            'schnittstelle': 'interface',
            'uuid': 'uuid',
            'id': 'id',
            'permission': 'permission',
            'berechtigung': 'permission',
            'nur': 'only',
            'admin': 'admin',
            'benutzer': 'user',
            'user': 'user',
            'erstellen': 'create',
            'ansicht': 'view',
            'template': 'template',
            'vorlage': 'template',
            'anzeigen': 'display',
            'show': 'show',
            'konvertieren': 'convert',
            'update': 'update',
            'änderung': 'change',
            'feld': 'field',
            'identifikator': 'identifier',
            'eingabe': 'input'
        }
        
        for de_word, en_word in replacements.items():
            text = text.replace(de_word, en_word)
        
        # Remover caracteres especiales excepto espacios
        import re
        text = re.sub(r'[^\w\s]', ' ', text)
        # Eliminar espacios múltiples
        text = ' '.join(text.split())
        
        return text
    
    def calculate_similarity(self, text1, text2):
        """Calcular similitud entre dos textos usando Jaccard mejorado"""
        if not text1 or not text2:
            return 0.0
        
        # Usar SequenceMatcher de difflib
        ratio = SequenceMatcher(None, text1, text2).ratio()
        
        # Método de Jaccard sobre palabras
        words1 = set(text1.split())
        words2 = set(text2.split())
        
        if not words1 and not words2:
            return 1.0
        if not words1 or not words2:
            return 0.0
        
        intersection = len(words1 & words2)
        union = len(words1 | words2)
        
        jaccard = intersection / union if union > 0 else 0.0
        
        # Método de n-gramas para capturar similitud de substrings
        bigrams1 = set()
        bigrams2 = set()
        
        for i in range(len(text1) - 1):
            bigrams1.add(text1[i:i+2])
        for i in range(len(text2) - 1):
            bigrams2.add(text2[i:i+2])
        
        if bigrams1 and bigrams2:
            bigram_intersection = len(bigrams1 & bigrams2)
            bigram_union = len(bigrams1 | bigrams2)
            bigram_sim = bigram_intersection / bigram_union if bigram_union > 0 else 0.0
        else:
            bigram_sim = 0.0
        
        # Ponderar las métricas: Jaccard es más importante para multiidioma
        return (jaccard * 0.5) + (ratio * 0.3) + (bigram_sim * 0.2)
    
    def map_tasks_to_issues(self, tasks, issues):
        """Mapear tasks a issues usando similitud semántica"""
        mappings = []
        
        for task in tasks:
            task_text = self.normalize_text(f"{task['title']} {task['description']}")
            
            best_issue = None
            best_score = 0.0
            
            for issue in issues:
                issue_text = self.normalize_text(f"{issue['title']} {issue['body']}")
                score = self.calculate_similarity(task_text, issue_text)
                
                if score > best_score:
                    best_score = score
                    best_issue = issue
            
            mapping = {
                'task': task,
                'issue': best_issue,
                'similarity_score': best_score,
                'is_mapped': best_score > 0.15  # Umbral reducido a 0.15
            }
            
            mappings.append(mapping)
        
        return mappings
    
    def calculate_stats(self, issues, mappings):
        """Calcular estadísticas"""
        total_issues = len(issues)
        closed_issues = len([i for i in issues if i['state'] == 'CLOSED'])
        
        total_tasks = len(mappings)
        mapped_tasks = len([m for m in mappings if m['is_mapped']])
        
        def percentage(num, den):
            return round((num / den * 100), 2) if den > 0 else 0
        
        return {
            'total_issues': total_issues,
            'closed_issues': closed_issues,
            'open_issues': total_issues - closed_issues,
            'completion_percentage': percentage(closed_issues, total_issues),
            'total_tasks': total_tasks,
            'mapped_tasks': mapped_tasks,
            'unmapped_tasks': total_tasks - mapped_tasks,
            'mapping_coverage': percentage(mapped_tasks, total_tasks)
        }
    
    def generate_excel(self, base_filename, mappings, stats):
        """Generar archivo Excel con los mapeos y estadísticas"""
        excel_file = f"{base_filename}.xlsx"
        
        if not OPENPYXL_AVAILABLE:
            print("❌ Error: openpyxl no está disponible")
            print("   Intenta: pip3 install openpyxl")
            return self.generate_csv_fallback(base_filename, mappings)
        
        try:
            wb = Workbook()
            
            # Hoja 1: Mapeos detallados
            ws_mappings = wb.active
            ws_mappings.title = "Task-Issue Mappings"
            self._create_mappings_sheet(ws_mappings, mappings)
            
            # Hoja 2: Resumen de estadísticas
            ws_summary = wb.create_sheet("Summary")
            self._create_summary_sheet(ws_summary, stats, mappings)
            
            # Hoja 3: Tasks sin mapeo
            unmapped = [m for m in mappings if not m['is_mapped']]
            if unmapped:
                ws_unmapped = wb.create_sheet("Unmapped Tasks")
                self._create_unmapped_sheet(ws_unmapped, unmapped)
            
            wb.save(excel_file)
            return excel_file
            
        except Exception as e:
            print(f"❌ Error al generar Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def _create_mappings_sheet(self, ws, mappings):
        """Crear hoja de mapeos detallados"""
        # Encabezados
        headers = [
            'Task ID',
            'Task Title',
            'Task Description',
            'Requirement ID',
            'Requirement Title',
            'Issue #',
            'Issue Title',
            'Issue Body',
            'Issue State',
            'Similarity Score',
            'Mapped'
        ]
        
        # Aplicar estilos a encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, mapping in enumerate(mappings, 2):
            task = mapping['task']
            issue = mapping['issue']
            score = round(mapping['similarity_score'], 3)
            mapped_str = '✓ Yes' if mapping['is_mapped'] else '✗ No'
            
            # Aplicar color de fondo según estado de mapeo
            row_fill = PatternFill(
                start_color="D4EDDA" if mapping['is_mapped'] else "F8D7DA",
                end_color="D4EDDA" if mapping['is_mapped'] else "F8D7DA",
                fill_type="solid"
            ) if mapping['is_mapped'] or score > 0 else None
            
            row_data = [
                task['id'],
                task['title'],
                task['description'],
                task['requirement_id'],
                task['requirement_title'],
                issue['number'] if issue else '—',
                issue['title'] if issue else '—',
                (issue['body'][:100] + '...') if issue and len(issue['body']) > 100 else (issue['body'] if issue else '—'),
                issue['state'] if issue else '—',
                score,
                mapped_str
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                if row_fill:
                    cell.fill = row_fill
                
                # Alineación especial para números
                if col_idx == 10:  # Similarity Score
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 11:  # Mapped
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar ancho de columnas
        column_widths = [12, 25, 35, 12, 25, 10, 25, 40, 12, 15, 12]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Congelar encabezado
        ws.freeze_panes = "A2"
    
    def _create_summary_sheet(self, ws, stats, mappings):
        """Crear hoja de resumen"""
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=12, color="FFFFFF")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "📊 TASK-ISSUE MAPPING REPORT"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 25
        
        row += 2
        
        # Sección: Issues
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "📈 ISSUE STATISTICS"
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        
        # Datos de issues
        issue_data = [
            ("Total Issues", stats['total_issues']),
            ("Closed Issues", stats['closed_issues']),
            ("Open Issues", stats['open_issues']),
            ("Completion Rate (%)", stats['completion_percentage']),
        ]
        
        for label, value in issue_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '0.00%' if isinstance(value, float) and value < 100 else '0'
            row += 1
        
        row += 1
        
        # Sección: Tasks
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "📋 TASK STATISTICS"
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        
        # Datos de tasks
        task_data = [
            ("Total Tasks", stats['total_tasks']),
            ("Mapped Tasks", stats['mapped_tasks']),
            ("Unmapped Tasks", stats['unmapped_tasks']),
            ("Mapping Coverage (%)", stats['mapping_coverage']),
        ]
        
        for label, value in task_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '0.00%' if isinstance(value, float) and value < 100 else '0'
            row += 1
        
        row += 2
        
        # Sección: Details
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "ℹ️  ADDITIONAL INFORMATION"
        cell.font = section_font
        cell.fill = section_fill
        row += 1
        
        # Información
        ws[f'A{row}'] = "Generated"
        ws[f'B{row}'] = datetime.now(timezone.utc).isoformat()
        row += 1
        
        ws[f'A{row}'] = "Algorithm"
        ws[f'B{row}'] = "Semantic Similarity (Jaccard + SequenceMatcher + Bigrams)"
        row += 1
        
        ws[f'A{row}'] = "Similarity Threshold"
        ws[f'B{row}'] = 0.15
        
        # Ajustar ancho
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    
    def _create_unmapped_sheet(self, ws, unmapped):
        """Crear hoja de tasks sin mapeo"""
        # Encabezados
        headers = [
            'Task ID',
            'Task Title',
            'Task Description',
            'Requirement ID',
            'Requirement Title'
        ]
        
        # Aplicar estilos a encabezados
        header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        for row_idx, mapping in enumerate(unmapped, 2):
            task = mapping['task']
            
            row_data = [
                task['id'],
                task['title'],
                task['description'],
                task['requirement_id'],
                task['requirement_title']
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Ajustar ancho de columnas
        column_widths = [12, 25, 35, 12, 25]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Congelar encabezado
        ws.freeze_panes = "A2"
    
    def generate_csv_fallback(self, base_filename, mappings):
        """Fallback a CSV si openpyxl no está disponible"""
        csv_file = f"{base_filename}.csv"
        
        headers = [
            'Task ID',
            'Task Title',
            'Task Description',
            'Requirement ID',
            'Requirement Title',
            'Issue #',
            'Issue Title',
            'Issue State',
            'Similarity Score',
            'Mapped'
        ]
        
        rows = []
        for mapping in mappings:
            task = mapping['task']
            issue = mapping['issue']
            score = round(mapping['similarity_score'], 3)
            mapped = '✓' if mapping['is_mapped'] else '✗'
            
            rows.append([
                task['id'],
                task['title'],
                task['description'],
                task['requirement_id'],
                task['requirement_title'],
                issue['number'] if issue else '—',
                issue['title'] if issue else '—',
                issue['state'] if issue else '—',
                score,
                mapped
            ])
        
        try:
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            return csv_file
        except Exception as e:
            print(f"❌ Error al generar CSV: {str(e)}")
            sys.exit(1)
    
    def generate_summary(self, base_filename, stats):
        """Generar archivo de resumen"""
        summary_file = f"{base_filename}-SUMMARY.txt"
        
        content = f"""╔════════════════════════════════════════════════════════════╗
║        TASK-ISSUE MAPPING REPORT SUMMARY                   ║
║                                                            ║
║  Análisis Semántico de Tasks vs GitHub Issues             ║
╚════════════════════════════════════════════════════════════╝

📊 ESTADÍSTICAS DE ISSUES
─────────────────────────────────────────────────────────────
Total de Issues:                  {stats['total_issues']}
Issues Cerrados (CLOSED):         {stats['closed_issues']}
Issues Abiertos (OPEN):           {stats['open_issues']}
Tasa de Completitud:              {stats['completion_percentage']}%

📋 ESTADÍSTICAS DE TASKS
─────────────────────────────────────────────────────────────
Total de Tasks:                   {stats['total_tasks']}
Tasks Mapeados a Issues:          {stats['mapped_tasks']}
Tasks sin Mapeo:                  {stats['unmapped_tasks']}
Cobertura de Mapeo:               {stats['mapping_coverage']}%

ℹ️  NOTAS IMPORTANTES
─────────────────────────────────────────────────────────────
• El mapeo se realiza usando similitud semántica de textos
• Soporta mezcla de alemán e inglés
• Umbral mínimo de similitud: 25%
• Revisa manualmente los mapeos para validar precisión

📁 ARCHIVOS GENERADOS
─────────────────────────────────────────────────────────────
CSV Report:                       {Path(base_filename).name}.csv
Summary:                          {Path(summary_file).name}

Timestamp:                        {datetime.now(timezone.utc).isoformat()}
"""
        
        try:
            with open(summary_file, 'w', encoding='utf-8') as f:
                f.write(content)
            return summary_file
        except Exception as e:
            print(f"❌ Error al generar resumen: {str(e)}")
            sys.exit(1)
    
    def run(self):
        """Ejecutar el mapeo completo"""
        print(f"🔍 Buscando archivos JSON en: {self.base_path}\n")
        
        # Validar archivos
        self.validate_files()
        
        print("✅ Archivos encontrados:")
        print(f"   - {self.requirements_file.name}")
        print(f"   - {self.board_issues_file.name}\n")
        
        # Cargar datos
        requirements = self.load_json(self.requirements_file)
        board_issues = self.load_json(self.board_issues_file)
        
        # Extraer tasks e issues
        tasks = self.extract_tasks(requirements)
        issues = self.extract_issues(board_issues)
        
        print("📋 Análisis de datos:")
        print(f"   - Tasks encontrados: {len(tasks)}")
        print(f"   - Issues encontrados: {len(issues)}\n")
        
        # Mapear tasks a issues
        print("🔗 Ejecutando mapeo semántico...")
        mappings = self.map_tasks_to_issues(tasks, issues)
        
        # Estadísticas
        stats = self.calculate_stats(issues, mappings)
        
        print("✅ Mapeo completado:")
        print(f"   - Tasks mapeados: {stats['mapped_tasks']}/{stats['total_tasks']}")
        print(f"   - Cobertura: {stats['mapping_coverage']}%")
        print(f"   - Issues cerrados: {stats['closed_issues']}/{stats['total_issues']}")
        print(f"   - Tasa de completitud: {stats['completion_percentage']}%\n")
        
        # Generar reportes
        print("📊 Generando reportes...")
        timestamp = datetime.now().strftime("%Y-%m-%d")
        base_filename = str(self.base_path / f"task-issue-mapping-{timestamp}")
        
        # Generar Excel (o CSV fallback si openpyxl no disponible)
        report_file = self.generate_excel(base_filename, mappings, stats)
        
        # Generar resumen
        summary_file = self.generate_summary(base_filename, stats)
        
        print("✨ Reportes generados:")
        print(f"   - {Path(report_file).name}")
        print(f"   - {Path(summary_file).name}\n")
        
        print("✅ ¡Proceso completado exitosamente!")


def main():
    if len(sys.argv) != 2:
        print("map-tasks-to-issues - Mapeo de Tasks a GitHub Issues")
        print("")
        print("Uso:")
        print("  python3 map_tasks_to_issues.py <nombre_repositorio>")
        print("")
        print("Ejemplo:")
        print("  python3 map_tasks_to_issues.py goetz-kundenportal-phoenix")
        print("")
        print("Descripción:")
        print("  Lee requirements.json y board-issues.json de:")
        print("  /Users/admin/dev/Reports/<nombre_repositorio>/")
        print("")
        print("  Genera un reporte CSV con el mapeo semántico entre")
        print("  tasks de requirements e issues de GitHub.")
        sys.exit(1)
    
    repo_name = sys.argv[1]
    mapper = TaskIssueMapper(repo_name)
    
    try:
        mapper.run()
    except KeyboardInterrupt:
        print("\n\n⚠️  Proceso cancelado por el usuario")
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Error inesperado: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
